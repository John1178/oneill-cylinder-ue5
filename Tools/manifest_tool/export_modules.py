"""
export_modules.py

Reads docs/module_list.xlsx, validates it, writes output/module_list.csv
for import into Unreal as a DataTable.

This script is the GATE between the human editing surface (Excel) and machine
data (CSV). If it writes a file, that file is guaranteed importable. If the
data is bad, it must refuse to write and say exactly which row is wrong.

Run:
    python export_modules.py
"""

import csv
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(HERE))
XLSX_PATH = os.path.join(PROJECT_ROOT, "docs", "module_list.xlsx")
OUT_PATH = os.path.join(HERE, "output", "module_list.csv")
SHEET_NAME = "Module List"


# ---------------------------------------------------------------------------
# What counts as valid
#
# These lists ARE the production standard. Anything not in them is rejected.
# Read from the current spreadsheet - update here if a category is ever added.
# ---------------------------------------------------------------------------

VALID_CATEGORIES = [
    "Building",
    "Structural Shell",
    "Habitat Surface",
    "Greenery & Terrain",
    "Infrastructure & Mechanical",
    "Debris & Wear Kit",
]

VALID_ZONES = ["Residential", "Industrial", "Service", "-"]
VALID_BELTS = ["A", "B", "All"]
VALID_ROTATION_MODES = ["Free", "AlignToRoad", "Fixed"]
VALID_SOURCES = ["Model", "Megascans", "Marketplace", "Scan"]

# Columns written to the CSV, in order. 'Name' MUST be first - UE requires it.
# Human-only columns (DisplayName, Status, Notes) are deliberately dropped:
# the DataTable UStruct has no fields for them.
CSV_COLUMNS = [
    "Name", "Category", "Zone", "Belt", "Mesh",
    "Weight", "RotationMode", "Clearance", "Source",
]


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

def read_rows(xlsx_path, sheet_name):
    """
    Open the workbook and return a list of dicts, one per module row.

    Each dict should also carry the spreadsheet row number, because every error
    message needs to tell the user WHICH row is wrong. Without that the script
    is useless on a 37-row sheet.

    Suggested shape:
        [{"_row": 2, "Name": "BLD_001", "Category": "Building", ...}, ...]

    Skip rows where Name is empty (blank rows at the bottom are normal).
    """
    # open the excel file and pick the sheet we want
    workbook = openpyxl.load_workbook(xlsx_path)
    ws = workbook[sheet_name]

    # step 1: get the column names out of row 1
    # after this, headers looks like ["Name", "DisplayName", "Category", ...]
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    rows = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num = row_num + 1

        module = {}
        for i in range(len(headers)):
            module[headers[i]] = row[i]
            
        rows.append(module)
        module["_row"] = row_num

    print(len(rows))
    return rows

# ---------------------------------------------------------------------------
# Validation
#
# Each check returns a list of human-readable error strings, or an empty list.
# Every message should name the row number and what was wrong.
#   good: "row 14: Weight 1.5 is outside 0.0-1.0"
#   bad:  "invalid weight"
# ---------------------------------------------------------------------------

def check_names(rows):
    """
    Name must exist, be unique, and look like PREFIX_NNN (e.g. BLD_001).

    Duplicates are the important one - two rows with the same ID means a
    manifest referencing it is ambiguous, and PCG picks non-deterministically.
    """
    
    # make an empty list called errors
    # make an empty list called seen_names
    errors = []
    seen_names = []

    for module in rows:
        name = module["Name"]

        if name is None or name == "":
            errors.append(f"row {module['_row']}: Name is empty")
            continue

        if name in seen_names:
            errors.append(f"row {module['_row']}: duplicate name")
        else:
            seen_names.append(name)

        parts = name.split("_")          

        if len(parts) != 2:
            errors.append(f"row {module['_row']}: bad format")
        elif parts[1].isdigit() == False:
            errors.append(f"row {module['_row']}: bad format")

    return errors

           
def check_enums(rows):
    """
    Category, Zone, Belt, RotationMode and Source must each be one of the
    VALID_* lists above.

    This is what catches typos like 'Buliding' - which would otherwise create a
    phantom category and silently drop the module out of its group.
    """
    # make an empty list called errors
    errors = []

    for module in rows:
        category = module["Category"]
        if category not in VALID_CATEGORIES:
            errors.append(f"row {module['_row']}: Category '{category}' is not valid")

        zone = module["Zone"]
        if zone not in VALID_ZONES:
            errors.append(f"row {module['_row']}: Zone '{zone}' is not valid")

        belt = module["Belt"]
        if belt not in VALID_BELTS:
            errors.append(f"row {module['_row']}: Belt '{belt}' is not valid")

        rotationmode = module["RotationMode"]
        if rotationmode not in VALID_ROTATION_MODES:
            errors.append(f"row {module['_row']}: RotationMode '{rotationmode}' is not valid")

        source = module["Source"]
        if source not in VALID_SOURCES:
            errors.append(f"row {module['_row']}: Source '{source}' is not valid")

    return errors



def check_zone_rules(rows):
    """
    Zone rules, which are project-specific rather than generic:

    - A Building MUST have a real zone (not '-'), otherwise the manifest's
        zone_ratios can never select it and it will never be placed.
    - A non-Building SHOULD have '-'. A tree tagged 'Residential' is a
        mistake, not a feature.
    """
    errors = []

    for module in rows:
        category = module["Category"]
        zone = module["Zone"]

        if category == "Building" and zone == "-":
            errors.append(f"row {module['_row']}: building has no zone")

        if category != "Building" and zone != "-":
            errors.append(f"row {module['_row']}: only building should have zone")

    return errors



def check_numbers(rows):
    """
    Weight must be a number in 0.0-1.0.
    Clearance must be a number >= 0.

    Note weight 0 is legal - it means 'retired, never place this' - so check
    the range, not truthiness.
    """

    errors = []

    for module in rows:
        weight = module["Weight"]
        clearance = module["Clearance"]

        if isinstance(weight, (int, float)) == False:
            errors.append(f"row {module['_row']}: Weight '{weight}' must be a number")
        elif weight < 0.0 or weight > 1.0:
            errors.append(f"row {module['_row']}: Weight '{weight}' outside 0.0 to 1.0")

        if isinstance(clearance, (int, float)) == False:
            errors.append(f"row {module['_row']}: Clearance '{clearance}' must be a number")
        elif clearance < 0:
            errors.append(f"row {module['_row']}: Clearance '{clearance}' must be a number >= 0")

    return errors

def check_mesh_paths(rows):
    """
    Mesh path checks.

    Blank is allowed - the asset may not exist yet. That is a WARNING, not an
    error: the module is simply not ready. If it is filled in, it should look
    like a UE content path (starts with /Game/).

    Return (errors, warnings) so the caller can treat them differently.
    """

    errors = []
    warning = []

    for module in rows:
        mesh = module["Mesh"]

        if not mesh:
            warning.append(f"row {module['_row']}: Mesh '{mesh}' not build yet")
            continue

        if isinstance(mesh, str) == False:
            errors.append(f"row {module['_row']}: Mesh '{mesh}' must be text")
        elif not mesh.startswith("/Game/"):
            errors.append(f"row {module['_row']}: Mesh '{mesh}' must be a UE path")

    return errors, warning 





# def validate(rows):
#     """
#     Run every check, collect all errors and warnings.

#     Do NOT stop at the first error - report everything at once so the user
#     fixes the sheet in one pass instead of re-running ten times.

#     Return (errors, warnings).
#     """
#     # TODO
#     raise NotImplementedError


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

# def write_csv(rows, out_path):
#     """
#     Write CSV_COLUMNS in order, 'Name' first.

#     Only called when there are zero errors. Warnings do not block the write.
#     """
#     # TODO
#     raise NotImplementedError


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    """
    1. read
    2. validate
    3. print warnings
    4. if errors: print them all and exit non-zero WITHOUT writing
    5. else write the CSV and report the count

    Exit code matters: a non-zero exit means this can be wired into a build
    step or pre-commit hook later and actually block bad data.
    """
    # TODO
    result = read_rows(XLSX_PATH, SHEET_NAME)
    print(len(result))
    print(result[1])

if __name__ == "__main__":
    main()
